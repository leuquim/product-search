"""
File management module for handling multiple Excel imports
"""
import os
import openpyxl
import pandas as pd
from typing import List, Dict, Any, Tuple, Optional
from tqdm import tqdm
import time
from datetime import datetime
from .database import MultiFileDatabase
from .config import CHUNK_SIZE, IMPORT_METHOD
from .fast_importer import FastExcelImporter


class FileManager:
    def __init__(self, database: MultiFileDatabase = None):
        self.database = database or MultiFileDatabase()
        self.chunk_size = CHUNK_SIZE

    def preview_file(self, file_path: str, rows: int = 10) -> Dict[str, Any]:
        """Preview an Excel file before importing"""
        try:
            # Get basic file info
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            filename = os.path.basename(file_path)
            
            # Preview data
            preview_data, headers = self.database.preview_excel_data(file_path, rows)
            
            # Suggest columns to index (those containing key terms)
            suggested_indexes = []
            index_keywords = ['assembly', 'part', 'number', 'description', 'name', 
                            'model', 'sku', 'id', 'code', 'serial']
            
            for header in headers:
                header_lower = header.lower()
                if any(keyword in header_lower for keyword in index_keywords):
                    suggested_indexes.append(header)
            
            # Get total row count estimate
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active
            total_rows = sheet.max_row - 1  # Exclude header
            workbook.close()
            
            return {
                "success": True,
                "filename": filename,
                "file_size_mb": round(file_size_mb, 2),
                "headers": headers,
                "suggested_indexes": suggested_indexes,
                "preview_data": preview_data,
                "total_rows": total_rows
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    def import_file(self, file_path: str, indexed_columns: List[str] = None,
                   progress_callback: callable = None, use_fast_import: bool = True) -> Dict[str, Any]:
        """Import an Excel file with selected indexed columns - now with fast import option"""
        start_time = time.time()
        filename = os.path.basename(file_path)
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        
        print(f"📁 Importing file: {filename} ({file_size_mb:.1f} MB)")
        
        # Register the file first
        file_id = self.database.register_file(
            filename=filename,
            file_size_mb=file_size_mb,
            indexed_columns=indexed_columns
        )
        
        # Determine import method
        import_method = IMPORT_METHOD
        if use_fast_import and import_method in ['auto', 'duckdb_native', 'pandas']:
            # Try fast import methods
            result = self._fast_import(file_path, file_id, indexed_columns, filename)
            if result.get("success"):
                # Update file stats and create indexes
                self.database.update_file_stats(file_id, result["rows_imported"])
                if indexed_columns:
                    print(f"🔍 Creating indexes on {len(indexed_columns)} columns...")
                    self.database.create_indexes(file_id, indexed_columns)
                self.database.conn.commit()
                
                result["file_id"] = file_id
                result["filename"] = filename
                result["file_size_mb"] = round(file_size_mb, 2)
                return result
            else:
                print("⚠️ Fast import failed, falling back to standard method...")
        
        # Fallback to original openpyxl method
        return self._standard_import(file_path, file_id, indexed_columns, 
                                    progress_callback, filename, file_size_mb)
    
    def _fast_import(self, file_path: str, file_id: int, 
                    indexed_columns: List[str], filename: str) -> Dict[str, Any]:
        """Use fast import methods"""
        try:
            # Create fast importer instance
            fast_importer = FastExcelImporter(self.database.conn)
            
            # Try auto import (will try multiple methods)
            result = fast_importer.auto_import(file_path, file_id, indexed_columns)
            
            if result.get("success"):
                print(f"✨ Fast import successful using {result.get('method')} method!")
                if 'rows_per_second' in result:
                    print(f"⚡ Performance: {result['rows_per_second']:,} rows/second")
            
            return result
            
        except Exception as e:
            print(f"❌ Fast import error: {e}")
            return {"success": False, "error": str(e)}
    
    def _standard_import(self, file_path: str, file_id: int, indexed_columns: List[str],
                        progress_callback: callable, filename: str, 
                        file_size_mb: float) -> Dict[str, Any]:
        """Original openpyxl-based import method"""
        start_time = time.time()
        
        try:
            print("📖 Using standard openpyxl import method...")
            
            # Read Excel file
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active
            
            # Get headers
            headers = []
            for cell in sheet[1]:
                if cell.value is not None and str(cell.value).strip():
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(f"Column_{len(headers)+1}")
            
            print(f"📋 Found {len(headers)} columns")
            
            # Register columns
            self.database.register_columns(file_id, headers, indexed_columns or [])
            
            # Import data in chunks
            rows_processed = 0
            chunk_data = []
            total_rows = sheet.max_row - 1
            
            # Create progress bar
            pbar = tqdm(total=total_rows, desc=f"Importing {filename}", unit="rows")
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                has_data = False
                
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        if value is not None:
                            row_dict[headers[idx]] = str(value).strip()
                            has_data = True
                        else:
                            row_dict[headers[idx]] = ""
                
                if has_data:
                    chunk_data.append(row_dict)
                    
                    if len(chunk_data) >= self.chunk_size:
                        # Insert chunk
                        inserted = self.database.insert_batch(chunk_data, file_id)
                        rows_processed += inserted
                        pbar.update(inserted)
                        
                        if progress_callback:
                            progress_callback({
                                "rows_processed": rows_processed,
                                "total_rows": total_rows,
                                "percentage": (rows_processed / total_rows) * 100
                            })
                        
                        chunk_data = []
            
            # Insert remaining data
            if chunk_data:
                inserted = self.database.insert_batch(chunk_data, file_id)
                rows_processed += inserted
                pbar.update(inserted)
            
            pbar.close()
            workbook.close()
            
            # Update file stats
            self.database.update_file_stats(file_id, rows_processed)
            
            # Create indexes on selected columns
            if indexed_columns:
                print(f"🔍 Creating indexes on {len(indexed_columns)} columns...")
                self.database.create_indexes(file_id, indexed_columns)
            
            # Commit changes
            self.database.conn.commit()
            
            duration = time.time() - start_time
            print(f"✅ Import completed: {rows_processed:,} rows in {duration:.2f}s")
            print(f"⚡ Speed: {rows_processed/duration:.0f} rows/second")
            
            return {
                "success": True,
                "method": "openpyxl",
                "file_id": file_id,
                "filename": filename,
                "rows_imported": rows_processed,
                "duration_seconds": round(duration, 2),
                "file_size_mb": round(file_size_mb, 2),
                "rows_per_second": round(rows_processed/duration)
            }
            
        except Exception as e:
            print(f"❌ Standard import failed: {e}")
            self.database.conn.rollback()
            return {
                "success": False,
                "error": str(e),
                "filename": filename
            }

    def delete_file(self, file_id: int) -> Dict[str, Any]:
        """Delete an imported file and its data"""
        try:
            # Get file info before deletion
            files = self.database.get_imported_files()
            file_info = next((f for f in files if f['file_id'] == file_id), None)
            
            if not file_info:
                return {
                    "success": False,
                    "error": "File not found"
                }
            
            # Delete the file
            if self.database.delete_file(file_id):
                return {
                    "success": True,
                    "filename": file_info['filename'],
                    "rows_deleted": file_info['row_count']
                }
            else:
                return {
                    "success": False,
                    "error": "Failed to delete file"
                }
                
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    def get_file_details(self, file_id: int) -> Dict[str, Any]:
        """Get detailed information about an imported file"""
        try:
            # Get file info
            files = self.database.get_imported_files()
            file_info = next((f for f in files if f['file_id'] == file_id), None)
            
            if not file_info:
                return {
                    "success": False,
                    "error": "File not found"
                }
            
            # Get column info
            columns = self.database.conn.execute("""
                SELECT column_name, is_indexed
                FROM file_columns
                WHERE file_id = ?
                ORDER BY column_name
            """, [file_id]).fetchall()
            
            column_info = [
                {"name": col[0], "indexed": col[1]}
                for col in columns
            ]
            
            # Get sample data
            sample_data, count = self.database.search("", [file_id], limit=5)
            
            return {
                "success": True,
                "file_info": file_info,
                "columns": column_info,
                "sample_data": sample_data,
                "indexed_columns": [c["name"] for c in column_info if c["indexed"]]
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    def update_indexes(self, file_id: int, indexed_columns: List[str]) -> Dict[str, Any]:
        """Update the indexed columns for a file"""
        try:
            # Update column index flags
            all_columns = self.database.conn.execute("""
                SELECT column_name FROM file_columns WHERE file_id = ?
            """, [file_id]).fetchall()
            
            for col_tuple in all_columns:
                col = col_tuple[0]
                is_indexed = col in indexed_columns
                
                self.database.conn.execute("""
                    UPDATE file_columns 
                    SET is_indexed = ?
                    WHERE file_id = ? AND column_name = ?
                """, [is_indexed, file_id, col])
            
            # Recreate indexes
            print(f"🔍 Updating indexes for file {file_id}...")
            
            # Drop old indexes (simplified - in production would track index names)
            # For now, we'll just create new ones
            self.database.create_indexes(file_id, indexed_columns)
            
            # Update imported_files table
            indexed_str = ','.join(indexed_columns)
            self.database.conn.execute("""
                UPDATE imported_files 
                SET indexed_columns = ?, updated_at = CURRENT_TIMESTAMP
                WHERE file_id = ?
            """, [indexed_str, file_id])
            
            self.database.conn.commit()
            
            return {
                "success": True,
                "file_id": file_id,
                "indexed_columns": indexed_columns
            }
            
        except Exception as e:
            self.database.conn.rollback()
            return {
                "success": False,
                "error": str(e)
            }

    def get_all_files(self) -> List[Dict[str, Any]]:
        """Get list of all imported files with statistics"""
        return self.database.get_imported_files()

    def export_search_results(self, query: str, file_ids: List[int] = None,
                            output_path: str = None) -> Dict[str, Any]:
        """Export search results to CSV"""
        try:
            # Perform search
            results, total_count = self.database.search(query, file_ids, limit=10000)
            
            if not results:
                return {
                    "success": False,
                    "error": "No results to export"
                }
            
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Generate output path if not provided
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"export_{timestamp}.csv"
            
            # Export to CSV
            df.to_csv(output_path, index=False)
            
            return {
                "success": True,
                "output_path": output_path,
                "rows_exported": len(results),
                "file_size_kb": round(os.path.getsize(output_path) / 1024, 2)
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }