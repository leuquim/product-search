import openpyxl
import pandas as pd
from typing import List, Dict, Any, Iterator
from tqdm import tqdm
import time
from .database import ProductDatabase
from .config import CHUNK_SIZE


class ExcelImporter:
    def __init__(self, excel_path: str, database: ProductDatabase = None):
        self.excel_path = excel_path
        self.database = database or ProductDatabase()
        self.chunk_size = CHUNK_SIZE

    def get_excel_info(self) -> Dict[str, Any]:
        """Get basic information about the Excel file"""
        try:
            workbook = openpyxl.load_workbook(self.excel_path, read_only=True)
            sheet = workbook.active
            
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Column_{col}")
            
            workbook.close()
            
            return {
                "total_rows": max_row - 1,  # Exclude header
                "total_columns": max_col,
                "headers": headers,
                "file_path": self.excel_path
            }
        except Exception as e:
            raise Exception(f"Failed to read Excel file info: {e}")

    def read_excel_chunks(self) -> Iterator[List[Dict[str, Any]]]:
        """Read Excel file in chunks for memory efficiency"""
        try:
            print("📖 Reading Excel file in chunks...")
            
            # For very large files, try to read in smaller chunks using openpyxl more efficiently
            workbook = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            
            # Get headers efficiently
            headers = []
            for idx, cell in enumerate(sheet[1]):
                if cell.value is not None and str(cell.value).strip():
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(f"Column_{idx+1}")
            print(f"📋 Found {len(headers)} columns")
            
            # Read data in chunks using sheet.iter_rows for better performance
            rows_processed = 0
            chunk_data = []
            
            # Skip header row and iterate through data
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Convert row to dictionary
                row_dict = {}
                has_data = False
                
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        header = headers[idx]
                        if value is not None:
                            row_dict[header] = str(value).strip()
                            has_data = True
                        else:
                            row_dict[header] = ""
                
                # Only add rows with some data
                if has_data:
                    chunk_data.append(row_dict)
                
                rows_processed += 1
                
                # Yield chunk when it reaches chunk_size
                if len(chunk_data) >= self.chunk_size:
                    print(f"📦 Processing chunk: {rows_processed:,} rows read, {len(chunk_data)} with data")
                    yield chunk_data
                    chunk_data = []
            
            # Yield remaining data
            if chunk_data:
                print(f"📦 Final chunk: {len(chunk_data)} rows")
                yield chunk_data
            
            workbook.close()
            print(f"✅ Finished reading {rows_processed:,} total rows")
                
        except Exception as e:
            raise Exception(f"Failed to read Excel file: {e}")

    def validate_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Validate and clean data before insertion"""
        cleaned_data = []
        
        for row in data:
            cleaned_row = {}
            for key, value in row.items():
                # Clean column names
                clean_key = str(key).strip() if key else "Unknown"
                
                # Clean values
                if pd.isna(value) or value is None:
                    clean_value = ""
                else:
                    clean_value = str(value).strip()
                
                cleaned_row[clean_key] = clean_value
            
            # Only add rows that have some non-empty values
            if any(v.strip() for v in cleaned_row.values() if v):
                cleaned_data.append(cleaned_row)
        
        return cleaned_data

    def import_excel(self, clear_existing: bool = False) -> Dict[str, Any]:
        """Import Excel file into database with progress tracking"""
        start_time = time.time()
        
        try:
            # Get file info
            excel_info = self.get_excel_info()
            print(f"Importing Excel file: {self.excel_path}")
            print(f"Total rows: {excel_info['total_rows']:,}")
            print(f"Total columns: {excel_info['total_columns']}")
            
            # Clear existing data if requested
            if clear_existing:
                self.database.clear_table()
                print("Cleared existing data")
            
            # Setup progress bar
            total_rows = excel_info['total_rows']
            progress_bar = tqdm(
                total=total_rows,
                desc="Importing",
                unit="rows",
                unit_scale=True
            )
            
            total_imported = 0
            total_processed = 0
            chunk_count = 0
            
            # Process chunks
            print(f"📊 Processing {total_rows:,} rows in chunks of {self.chunk_size:,}...")
            
            for chunk_data in self.read_excel_chunks():
                chunk_count += 1
                
                if not chunk_data:  # Skip empty chunks
                    continue
                
                # Validate and clean data
                cleaned_data = self.validate_data(chunk_data)
                
                if cleaned_data:
                    # Insert into database
                    inserted_count = self.database.insert_batch(cleaned_data)
                    total_imported += inserted_count
                
                # Update progress based on raw chunk size (including empty rows)
                total_processed += len(chunk_data)
                progress_bar.update(len(chunk_data))
                progress_bar.set_postfix({
                    'Chunk': chunk_count,
                    'Imported': f"{total_imported:,}",
                    'Valid': f"{len(cleaned_data):,}"
                })
                
                # Debug: print first chunk info
                if chunk_count == 1:
                    print(f"\n💫 First chunk processed: {len(chunk_data)} rows, {len(cleaned_data)} valid rows imported")
            
            progress_bar.close()
            
            # Calculate statistics
            end_time = time.time()
            duration = end_time - start_time
            rows_per_second = total_imported / duration if duration > 0 else 0
            
            result = {
                "success": True,
                "total_imported": total_imported,
                "chunks_processed": chunk_count,
                "duration_seconds": round(duration, 2),
                "rows_per_second": round(rows_per_second, 0),
                "file_info": excel_info
            }
            
            print(f"\nImport completed successfully!")
            print(f"Imported: {total_imported:,} rows")
            print(f"Duration: {duration:.2f} seconds")
            print(f"Speed: {rows_per_second:.0f} rows/second")
            
            return result
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "total_imported": 0
            }

    def preview_excel(self, num_rows: int = 5) -> List[Dict[str, Any]]:
        """Preview first few rows of Excel file"""
        try:
            df = pd.read_excel(
                self.excel_path,
                nrows=num_rows,
                dtype=str,
                na_filter=False
            )
            
            return df.to_dict('records')
            
        except Exception as e:
            raise Exception(f"Failed to preview Excel file: {e}")

    def get_column_sample(self, column_name: str, sample_size: int = 10) -> List[str]:
        """Get sample values from a specific column"""
        try:
            df = pd.read_excel(
                self.excel_path,
                usecols=[column_name],
                nrows=1000,  # Read first 1000 rows to sample from
                dtype=str,
                na_filter=False
            )
            
            # Get unique non-empty values
            unique_values = df[column_name].dropna().unique()
            sample_values = [v for v in unique_values if str(v).strip()][:sample_size]
            
            return sample_values
            
        except Exception as e:
            raise Exception(f"Failed to get column sample: {e}")


def import_excel_file(excel_path: str, clear_existing: bool = False) -> Dict[str, Any]:
    """Convenience function to import an Excel file"""
    with ProductDatabase() as db:
        importer = ExcelImporter(excel_path, db)
        return importer.import_excel(clear_existing)


def preview_excel_file(excel_path: str, num_rows: int = 5) -> List[Dict[str, Any]]:
    """Convenience function to preview an Excel file"""
    importer = ExcelImporter(excel_path)
    return importer.preview_excel(num_rows)